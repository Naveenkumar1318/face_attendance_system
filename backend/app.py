# Flask Imports
from flask import (Flask, render_template, request, redirect, url_for,flash, session, make_response, send_file)
# Standard Library Imports
import os
import shutil
import base64
import secrets
import atexit
from datetime import datetime, timedelta, time
from contextlib import contextmanager
from io import BytesIO
# Third-Party Libraries
import cv2
import numpy as np
import face_recognition
import pandas as pd
from geopy.distance import geodesic
from apscheduler.schedulers.background import BackgroundScheduler
import pymysql
import pymysql.cursors
import mysql.connector
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask_mail import Mail, Message

from io import BytesIO
from flask import send_file, flash, redirect, url_for
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl import Workbook
from datetime import date  # <-


# ---  Flask App Initialization ---
app = Flask(__name__)
app.secret_key = 'your_secret_key'
# --- Upload folder configuration ---
UPLOAD_FOLDER = os.path.join("static", "student_photos")
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)  # ensure folder exists

# ---- Mail Configuration ----
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'naveenn13032004@gmail.com'
app.config['MAIL_PASSWORD'] = 'vagr plhz jqgz wnob'
mail = Mail(app)

# ---- Database Connection & Session ----
def get_db_connection():
    conn = pymysql.connect(
        host='localhost',
        user='root',
        password='',
        db='ace_attendance_system',
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )
    # Keep sessions short and reduce blocking
    with conn.cursor() as cur:
        # MySQL 8 variable names
        try:
            cur.execute("SET SESSION transaction_isolation='READ-COMMITTED'")
        except Exception:
            pass
        try:
            cur.execute("SET SESSION innodb_lock_wait_timeout=10")  # default ~50
        except Exception:
            pass
    return conn

@contextmanager
def db_session(commit_on_success=True):
    """Context manager to ensure commit/rollback/finally close every time.
    Usage:
        with db_session() as cur:
            cur.execute(...)
    """
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        yield cur
        if commit_on_success:
            conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        try:
            cur.close()
        finally:
            conn.close()

# --- Helper Functions ---
def capture_image_from_base64(data):
    header, encoded = data.split(",")
    img_data = base64.b64decode(encoded)
    np_arr = np.frombuffer(img_data, np.uint8)
    return cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

#--- Face Encoding & Comparison ---
def encode_face_from_frame(frame):
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_frame)
    if not face_locations:
        return None, None
    return face_recognition.face_encodings(rgb_frame, face_locations)[0], face_locations[0]

#--- Serialize/Deserialize Encoding ---
def serialize_encoding(encoding):
    return ",".join(map(str, encoding))


def deserialize_encoding(string):
    return np.array([float(val) for val in string.split(",")])

#--- Compare Faces ---
def compare_faces(known_encoding, current_encoding, tolerance=0.45):
    distance = face_recognition.face_distance([known_encoding], current_encoding)[0]
    return distance <= tolerance

# --- Save Image ---
def save_image(frame, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    cv2.imwrite(path, frame)
# --- Save Attendance Image ---
# --- Save Attendance Image ---
def save_attendance_image(frame, department, year, register_number):
    """
    Save attendance image as:
    static/attendance_images/<Department>/<Year>/<Date>/<RegisterNo>.jpg
    """
    today_date = datetime.now().strftime("%Y-%m-%d")
    folder_path = os.path.join("static", "attendance_images", department, year, today_date)
    os.makedirs(folder_path, exist_ok=True)

    file_name = f"{register_number}.jpg"
    file_path = os.path.join(folder_path, file_name)

    cv2.imwrite(file_path, frame)

    return f"attendance_images/{department}/{year}/{today_date}/{register_number}.jpg"

# --- Time Slot Check ---
def get_current_slot():
    now = datetime.now().time()
    if now <= time(10, 0):
        return "morning"
    elif now >= time(14, 0):
        return "afternoon"
    else:
        return None

# ---- Location check ----

def is_within_college_location(lat, lon, radius_km=0.5):
    """
    Check if the given lat, lon is within 'radius_km' kilometers of the college.
    """
    college_lat = 12.7179
    college_lon = 77.8700
    distance_km = geodesic((lat, lon), (college_lat, college_lon)).km
    return distance_km <= radius_km


# --- Routes ---
@app.route('/')
def index():
    return render_template('index.html')

# --- Admin Login ---
@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        try:
            with db_session(commit_on_success=False) as cur:
                cur.execute("SELECT * FROM admin WHERE username = %s", (username,))
                admin = cur.fetchone()
        except Exception as e:
            flash(f"DB error: {e}")
            admin = None

        if not admin:
            flash("❌ Invalid username")
        elif not check_password_hash(admin['password'], password):
            flash("❌ Invalid password")
        else:
            session['admin_id'] = admin['id']
            session['admin_username'] = admin['username']
            session['admin_logged_in'] = True
            return redirect('/admin_dashboard')

    return render_template('admin_login.html')


# --- Admin Dashboard ---
@app.route('/admin_dashboard', methods=['GET', 'POST'])
def admin_dashboard():
    show = request.args.get('show', 'present')  # present or absent tab

    # --- Filters ---
    filter_department = request.form.get('filter_department', '')
    filter_year = request.form.get('filter_year', '')
    slot = request.form.get('slot', '')  # morning/afternoon
    student_search = request.form.get('student_search', '')  # for Manage Students

    # --- Pagination settings ---
    page_students = int(request.args.get('page_students', 1))
    per_page_students = 10
    offset_students = (page_students - 1) * per_page_students

    page_attendance = int(request.args.get('page_attendance', 1))
    per_page_attendance = 10
    offset_attendance = (page_attendance - 1) * per_page_attendance

    # --- Manage Students search ---
    students = []
    total_students = 0
    total_pages_students = 0

    if student_search:  # Only search by name or register_number
        students_query = "SELECT * FROM students WHERE register_number LIKE %s OR name LIKE %s"
        count_query = "SELECT COUNT(*) as count FROM students WHERE register_number LIKE %s OR name LIKE %s"
        params = [f"%{student_search}%", f"%{student_search}%"]
        count_params = [f"%{student_search}%", f"%{student_search}%"]

        students_query += " ORDER BY name ASC LIMIT %s OFFSET %s"
        params.extend([per_page_students, offset_students])

        try:
            with db_session(commit_on_success=False) as cur:
                cur.execute(students_query, params)
                students = cur.fetchall()
                cur.execute(count_query, count_params)
                total_students = cur.fetchone()['count']
        except Exception as e:
            flash(f"DB error (students search): {e}")
            students = []
            total_students = 0

        total_pages_students = (total_students + per_page_students - 1) // per_page_students

    # Normalize student photo paths
    for student in students:
        if student.get('photo'):
            student['photo'] = student['photo'].replace("\\", "/")

    # --- Attendance filtering ---
    attendance_records = []
    total_attendance_records = 0
    message = None

    try:
        with db_session(commit_on_success=False) as cur:
            if show == 'present':
                # Base query for today’s attendance
                attendance_query = """
                    SELECT a.id, a.student_id, s.register_number, s.name, s.department, s.year,
                           a.date, a.time, a.latitude, a.longitude, a.live_image
                    FROM attendance a
                    JOIN students s ON a.student_id = s.id
                    WHERE DATE(a.date) = CURDATE()
                """
                count_query_attendance = """
                    SELECT COUNT(*) as count
                    FROM attendance a
                    JOIN students s ON a.student_id = s.id
                    WHERE DATE(a.date) = CURDATE()
                """
                att_params = []
                count_params = []

                # Apply department/year filters
                if filter_department:
                    attendance_query += " AND s.department=%s"
                    count_query_attendance += " AND s.department=%s"
                    att_params.append(filter_department)
                    count_params.append(filter_department)

                if filter_year:
                    attendance_query += " AND s.year=%s"
                    count_query_attendance += " AND s.year=%s"
                    att_params.append(filter_year)
                    count_params.append(filter_year)

                # Apply slot filters
                if slot == 'morning':
                    attendance_query += " AND TIME(a.time) BETWEEN '08:00:00' AND '12:00:00'"
                    count_query_attendance += " AND TIME(a.time) BETWEEN '08:00:00' AND '12:00:00'"
                elif slot == 'afternoon':
                    attendance_query += " AND TIME(a.time) BETWEEN '13:00:00' AND '17:00:00'"
                    count_query_attendance += " AND TIME(a.time) BETWEEN '13:00:00' AND '17:00:00'"

                # Pagination
                attendance_query += " ORDER BY a.date DESC, a.time DESC LIMIT %s OFFSET %s"
                att_params.extend([per_page_attendance, offset_attendance])

                cur.execute(attendance_query, att_params)
                attendance_records = cur.fetchall()

                cur.execute(count_query_attendance, count_params)
                total_attendance_records = cur.fetchone()['count']

                message = "✅ No students marked as present today." if not attendance_records else None

            elif show == 'absent':
                # Absent students query (slot-aware)
                absent_query = """
                    SELECT s.register_number, s.name, s.department, s.year
                    FROM students s
                    WHERE s.id NOT IN (
                        SELECT a.student_id
                        FROM attendance a
                        WHERE DATE(a.date) = CURDATE()
                """

                # --- Slot filter ---
                if slot == 'morning':
                    absent_query += " AND TIME(a.time) BETWEEN '08:00:00' AND '12:00:00'"
                elif slot == 'afternoon':
                    absent_query += " AND TIME(a.time) BETWEEN '13:00:00' AND '17:00:00'"

                absent_query += ")"

                # Department/year filters
                absent_params = []
                count_query_absent = "SELECT COUNT(*) as count FROM students s WHERE 1=1"
                count_params = []

                if filter_department:
                    absent_query += " AND s.department=%s"
                    count_query_absent += " AND s.department=%s"
                    absent_params.append(filter_department)
                    count_params.append(filter_department)

                if filter_year:
                    absent_query += " AND s.year=%s"
                    count_query_absent += " AND s.year=%s"
                    absent_params.append(filter_year)
                    count_params.append(filter_year)

                absent_query += " ORDER BY s.name ASC LIMIT %s OFFSET %s"
                absent_params.extend([per_page_attendance, offset_attendance])

                cur.execute(absent_query, absent_params)
                attendance_records = cur.fetchall()

                cur.execute(count_query_absent, count_params)
                total_attendance_records = cur.fetchone()['count']

                message = "❌ No absent students today." if not attendance_records else None

    except Exception as e:
        flash(f"DB error (attendance): {e}")
        attendance_records = []
        total_attendance_records = 0

    total_pages_attendance = (total_attendance_records + per_page_attendance - 1) // per_page_attendance

    return render_template('admin_dashboard.html',
                           students=students,
                           student_search=student_search,
                           filter_department=filter_department,
                           filter_year=filter_year,
                           page_students=page_students,
                           total_pages_students=total_pages_students,
                           attendance_records=attendance_records,
                           page_attendance=page_attendance,
                           total_pages_attendance=total_pages_attendance,
                           show=show,
                           slot=slot,
                           message=message)

# --- Admin Dashboard  curd operation---
@app.route('/register_student', methods=['POST'])
def register_student():
    name = request.form['name']
    register_number = request.form['register_number']
    department = request.form['department']
    year = request.form['year']
    course_type = request.form['course_type']
    batch = request.form['batch']
    dob = request.form['dob']
    mobile = request.form['mobile']
    email = request.form['email']
    password = request.form['password']
    photo = request.files.get('photo')

    try:
        hashed_pw = generate_password_hash(password)

        photo_path_db = None
        face_encoding_str = None

        if photo and photo.filename != "":
            # Save photo in structured folder: static/student_photos/<Department>/<Year>/
            folder_path = os.path.join(app.config['UPLOAD_FOLDER'], department, year)
            os.makedirs(folder_path, exist_ok=True)
            photo_filename = f"{register_number}.jpg"
            save_path = os.path.join(folder_path, photo_filename)
            photo.save(save_path)

            # Generate face encoding
            frame = cv2.imread(save_path)
            if frame is not None:
                encoding, _ = encode_face_from_frame(frame)
                if encoding is not None:
                    face_encoding_str = serialize_encoding(encoding)

            # Store relative path in DB
            photo_path_db = os.path.relpath(save_path, "static").replace("\\", "/")

        # Insert student into DB
        with db_session() as cur:
            sql = """
                INSERT INTO students 
                (name, register_number, department, year, course_type, batch, dob, mobile, email, password, photo, face_encoding)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            val = (name, register_number, department, year, course_type,
                   batch, dob, mobile, email, hashed_pw, photo_path_db, face_encoding_str)
            cur.execute(sql, val)

        flash("✅ Student registered successfully with face encoding!")
        return redirect(url_for('admin_dashboard'))

    except Exception as e:
        flash(f"❌ Registration failed: {e}")
        return redirect(url_for('admin_dashboard'))
# ---  admin Update student ---
@app.route('/update_student/<int:id>', methods=['POST'])
def update_student(id):
    # Fetch student from DB
    with db_session() as cur:
        cur.execute("SELECT * FROM students WHERE id=%s", (id,))
        student = cur.fetchone()

    if not student:
        flash("❌ Student not found!", "error")
        return redirect(url_for('admin_dashboard'))

    # Get form data
    name = request.form.get('name')
    register_number = request.form.get('register_number')
    department = request.form.get('department')
    year = request.form.get('year')
    course_type = request.form.get('course_type')
    batch = request.form.get('batch')
    dob = request.form.get('dob')
    mobile = request.form.get('mobile')
    email = request.form.get('email')
    new_password = request.form.get('new_password')
    photo_file = request.files.get('photo')

    # Handle photo
    photo_filename = student['photo']  # default to existing photo

    if photo_file and photo_file.filename != '':
        # Remove old photo if exists
        if student['photo']:
            old_photo_path = os.path.join("static", student['photo'])
            if os.path.exists(old_photo_path):
                try:
                    os.remove(old_photo_path)
                except Exception:
                    pass

        # Save new photo in structured folder
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], department, year)
        os.makedirs(folder_path, exist_ok=True)
        filename = f"{register_number}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(photo_file.filename)}"
        file_path = os.path.join(folder_path, filename)
        photo_file.save(file_path)
        photo_filename = os.path.relpath(file_path, "static").replace("\\", "/")

        # Update face encoding
        frame = cv2.imread(file_path)
        face_encoding_str = None
        if frame is not None:
            encoding, _ = encode_face_from_frame(frame)
            if encoding is not None:
                face_encoding_str = serialize_encoding(encoding)
    else:
        face_encoding_str = student['face_encoding']

    # Handle password
    password_hash = student['password']
    if new_password and new_password.strip() != '':
        password_hash = generate_password_hash(new_password)

    # Update student in DB
    with db_session() as cur:
        cur.execute("""
            UPDATE students
            SET name=%s, register_number=%s, department=%s, year=%s, course_type=%s, batch=%s,
                dob=%s, mobile=%s, email=%s, password=%s, photo=%s, face_encoding=%s, last_updated=NOW()
            WHERE id=%s
        """, (name, register_number, department, year, course_type, batch,
              dob, mobile, email, password_hash, photo_filename, face_encoding_str, id))

    flash("✅ Student updated successfully!")
    return redirect(url_for('admin_dashboard'))

# ---- admin Delete student ----
@app.route('/delete_student/<int:id>', methods=['POST'])
def delete_student(id):
    if not session.get('admin_logged_in'):
        return redirect('/admin_login')

    try:
        with db_session(commit_on_success=False) as cur:
            cur.execute("SELECT photo FROM students WHERE id = %s", (id,))
            result = cur.fetchone()
        if result and result.get('photo'):
            photo_path = os.path.join('static', result['photo'])
            if os.path.exists(photo_path):
                os.remove(photo_path)
        with db_session() as cur:
            cur.execute("DELETE FROM students WHERE id = %s", (id,))
        flash("Student deleted successfully!", "success")
    except Exception as e:
        flash(f"❌ Delete failed: {e}")
    return redirect(url_for('admin_dashboard'))


# --- Format DataFrame ---
def format_attendance_dataframe(rows):
    df = pd.DataFrame(rows, columns=['date', 'time', 'slot', 'register_number', 'name', 'department', 'year'])
    
    # Format date
    df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.date
    
    # Convert time from timedelta or string
    def parse_time(value):
        try:
            return pd.to_datetime(value).time()
        except:
            try:
                td = pd.to_timedelta(value)
                return (pd.Timestamp('00:00:00') + td).time()
            except:
                return None
    
    df['time'] = df['time'].apply(parse_time)
    
    # Convert to string for Excel display
    df['date'] = df['date'].astype(str)
    df['time'] = df['time'].astype(str)
    
    return df

# --- Auto-adjust column widths ---
def auto_adjust_excel_columns(writer, sheet_name, dataframe):
    worksheet = writer.sheets[sheet_name]
    for i, column in enumerate(dataframe.columns, 1):
        max_length = max(
            dataframe[column].astype(str).map(len).max(),
            len(column)
        ) + 2
        worksheet.column_dimensions[get_column_letter(i)].width = max_length

# ---------- Helper: Convert rows to DataFrame ----------
def format_attendance_dataframe(rows, columns):
    df = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
    return df

# ---------- Helper: Auto-adjust Excel columns ----------
def auto_adjust_excel_columns(ws):
    for i, col in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col) + 2
        ws.column_dimensions[get_column_letter(i)].width = max(max_len, 15)

# ---------- Helper: Add Excel Header ----------
def add_excel_header(ws, title_lines):
    for _ in range(len(title_lines)):
        ws.insert_rows(1)
    for i, line in enumerate(title_lines, start=1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)
        cell = ws.cell(row=i, column=1)
        cell.value = line
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

# ---------- Helper: Write DataFrame or 'No data' ----------
def write_df_to_sheet(writer, sheet_name, df, title_lines):
    # Create or get sheet
    if sheet_name in writer.book.sheetnames:
        ws = writer.sheets[sheet_name]
    else:
        ws = writer.book.create_sheet(sheet_name)

    if df.empty:
        ws["A5"] = "No data available for this slot"
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=max(5, len(title_lines)))
        ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A5"].font = Font(bold=True, color="000000")
        ws["A5"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    else:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        auto_adjust_excel_columns(ws)

    add_excel_header(ws, title_lines)

# ==================== Export by Date (Morning & Afternoon) ====================
@app.route('/export_by_date_excel', methods=['POST'])
def export_by_date_excel():
    export_date = request.form.get('export_date')
    columns = ['Date', 'Time', 'Register Number', 'Name', 'Department', 'Year']

    try:
        with db_session(commit_on_success=False) as cur:
            # Morning
            query_morning = """
                SELECT a.date, TIME_FORMAT(a.time,'%%H:%%i:%%s') AS time,
                       s.register_number, s.name, s.department, s.year
                FROM attendance a
                JOIN students s ON a.student_id = s.id
                WHERE a.date=%s AND a.slot='morning'
                ORDER BY a.time ASC, s.register_number ASC
            """
            cur.execute(query_morning, (export_date,))
            morning_rows = cur.fetchall()

            # Afternoon
            query_afternoon = """
                SELECT a.date, TIME_FORMAT(a.time,'%%H:%%i:%%s') AS time,
                       s.register_number, s.name, s.department, s.year
                FROM attendance a
                JOIN students s ON a.student_id = s.id
                WHERE a.date=%s AND a.slot='afternoon'
                ORDER BY a.time ASC, s.register_number ASC
            """
            cur.execute(query_afternoon, (export_date,))
            afternoon_rows = cur.fetchall()

    except Exception as e:
        flash(f"❌ Export failed: {e}")
        return redirect(url_for('admin_dashboard'))

    df_morning = format_attendance_dataframe(morning_rows, columns)
    df_afternoon = format_attendance_dataframe(afternoon_rows, columns)

    if df_morning.empty and df_afternoon.empty:
        flash("⚠️ No attendance records found for this date.")
        return redirect(url_for('admin_dashboard'))

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        title_lines = [
            "ADHIYAAMAAN COLLEGE OF ENGINEERING, HOSUR",
            f"Attendance Records - {export_date}"
        ]
        write_df_to_sheet(writer, "Morning Attendance", df_morning, title_lines)
        write_df_to_sheet(writer, "Afternoon Attendance", df_afternoon, title_lines)

    output.seek(0)
    return send_file(
        output,
        download_name=f"attendance_{export_date}.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ==================== Export by Register Number (Morning & Afternoon) ====================
@app.route('/export_by_register_excel', methods=['POST'])
def export_by_register_excel():
    reg = request.form.get('register_number')
    columns = ['Date', 'Time', 'Register Number', 'Name', 'Department', 'Year']

    try:
        with db_session(commit_on_success=False) as cur:
            # Morning
            query_morning = """
                SELECT a.date, TIME_FORMAT(a.time,'%%H:%%i:%%s') AS time,
                       s.register_number, s.name, s.department, s.year
                FROM attendance a
                JOIN students s ON a.student_id = s.id
                WHERE s.register_number=%s AND a.slot='morning'
                ORDER BY a.date DESC, a.time ASC
            """
            cur.execute(query_morning, (reg,))
            morning_rows = cur.fetchall()

            # Afternoon
            query_afternoon = """
                SELECT a.date, TIME_FORMAT(a.time,'%%H:%%i:%%s') AS time,
                       s.register_number, s.name, s.department, s.year
                FROM attendance a
                JOIN students s ON a.student_id = s.id
                WHERE s.register_number=%s AND a.slot='afternoon'
                ORDER BY a.date DESC, a.time ASC
            """
            cur.execute(query_afternoon, (reg,))
            afternoon_rows = cur.fetchall()

    except Exception as e:
        flash(f"❌ Export failed: {e}")
        return redirect(url_for('admin_dashboard'))

    df_morning = format_attendance_dataframe(morning_rows, columns)
    df_afternoon = format_attendance_dataframe(afternoon_rows, columns)

    if df_morning.empty and df_afternoon.empty:
        flash("⚠️ No attendance records found for this student.")
        return redirect(url_for('admin_dashboard'))

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        title_lines = [
            "ADHIYAAMAAN COLLEGE OF ENGINEERING, HOSUR",
            f"Attendance Records - {reg}"
        ]
        write_df_to_sheet(writer, "Morning Attendance", df_morning, title_lines)
        write_df_to_sheet(writer, "Afternoon Attendance", df_afternoon, title_lines)

    output.seek(0)
    return send_file(
        output,
        download_name=f"attendance_{reg}.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



#-------------------EXPORT all ATTENDANCE-------------------------------------

# ---------- Helper: Convert rows to DataFrame ----------
def format_attendance_dataframe(rows, columns=None):
    df = pd.DataFrame(rows)
    if columns and not df.empty:
        df.columns = columns
    return df

# ---------- Helper: Auto-adjust Excel columns ----------
def auto_adjust_excel_columns(ws):
    for i, col in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col) + 2
        ws.column_dimensions[get_column_letter(i)].width = max(max_len, 15)

# ---------- Helper: Add Excel Header ----------
def add_excel_header(ws, title_lines):
    for _ in range(len(title_lines)):
        ws.insert_rows(1)
    for i, line in enumerate(title_lines, start=1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ws.max_column)
        cell = ws.cell(row=i, column=1)
        cell.value = line
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

# ---------- Helper: Write DataFrame or 'No data' ----------
def write_df_to_sheet(writer, sheet_name, df, title_lines):
    # Create or get sheet
    if sheet_name in writer.book.sheetnames:
        ws = writer.sheets[sheet_name]
    else:
        ws = writer.book.create_sheet(sheet_name)

    # Determine number of columns
    num_cols = len(df.columns) if not df.empty else max(5, len(title_lines))

    if df.empty:
        # Message for empty sheet (centered horizontally & vertically)
        ws["A5"] = "No data available for this slot"
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=num_cols)
        ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A5"].font = Font(bold=True, color="000000")
        ws["A5"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    else:
        # Write DataFrame to sheet
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        auto_adjust_excel_columns(ws)

    # Add college header + title
    add_excel_header(ws, title_lines)

    # Ensure all columns wide enough to avoid cut-off
    for i in range(1, num_cols + 1):
        if ws.column_dimensions[get_column_letter(i)].width < 25:
            ws.column_dimensions[get_column_letter(i)].width = 25



# ======================================================
#          DOWNLOAD TODAY PRESENT
# ======================================================
@app.route('/download_present', methods=['POST'])
def download_present():
    filter_dept = request.form.get('filter_dept', '')
    filter_year = request.form.get('filter_year', '')
    today = date.today()
    columns = ['Date', 'Time', 'Register Number', 'Name', 'Department', 'Year']

    try:
        with db_session(commit_on_success=False) as cur:
            # Morning
            query_morning = """
                SELECT a.date, TIME_FORMAT(a.time, '%%H:%%i:%%s'),
                       s.register_number, s.name, s.department, s.year
                FROM attendance a
                JOIN students s ON a.student_id = s.id
                WHERE a.date=%s AND a.slot='morning'
            """
            params = [today]
            if filter_dept:
                query_morning += " AND s.department=%s"
                params.append(filter_dept)
            if filter_year:
                query_morning += " AND s.year=%s"
                params.append(filter_year)
            cur.execute(query_morning, params)
            morning_rows = cur.fetchall()

            # Afternoon
            query_afternoon = query_morning.replace("a.slot='morning'", "a.slot='afternoon'")
            cur.execute(query_afternoon, params)
            afternoon_rows = cur.fetchall()
    except Exception as e:
        flash(f"❌ Download failed: {e}")
        return redirect(url_for('admin_dashboard'))

    df_morning = format_attendance_dataframe(morning_rows, columns)
    df_afternoon = format_attendance_dataframe(afternoon_rows, columns)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        title_lines = [
            "ADHIYAAMAAN COLLEGE OF ENGINEERING, HOSUR",
            f"Today's Present List - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        write_df_to_sheet(writer, "Morning Present", df_morning, title_lines)
        write_df_to_sheet(writer, "Afternoon Present", df_afternoon, title_lines)

    output.seek(0)
    return send_file(
        output,
        download_name=f'Present_Students_{today}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ======================================================
#          DOWNLOAD TODAY ABSENT
# ======================================================
@app.route('/download_absent', methods=['POST'])
def download_absent():
    filter_dept = request.form.get('filter_dept', '')
    filter_year = request.form.get('filter_year', '')
    today = date.today()
    columns = ['Register Number', 'Name', 'Department', 'Year', 'Mobile']

    try:
        with db_session(commit_on_success=False) as cur:
            # Morning Absent
            query_morning = """
                SELECT s.register_number, s.name, s.department, s.year, s.mobile
                FROM students s
                WHERE s.id NOT IN (SELECT student_id FROM attendance WHERE date=%s AND slot='morning')
            """
            params = [today]
            if filter_dept:
                query_morning += " AND s.department=%s"
                params.append(filter_dept)
            if filter_year:
                query_morning += " AND s.year=%s"
                params.append(filter_year)
            cur.execute(query_morning, params)
            morning_rows = cur.fetchall()

            # Afternoon Absent
            query_afternoon = query_morning.replace("slot='morning'", "slot='afternoon'")
            cur.execute(query_afternoon, params)
            afternoon_rows = cur.fetchall()
    except Exception as e:
        flash(f"❌ Download failed: {e}")
        return redirect(url_for('admin_dashboard'))

    df_morning = format_attendance_dataframe(morning_rows, columns)
    df_afternoon = format_attendance_dataframe(afternoon_rows, columns)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        title_lines = [
            "ADHIYAAMAAN COLLEGE OF ENGINEERING, HOSUR",
            f"Today's Absent List - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        write_df_to_sheet(writer, "Morning Absent", df_morning, title_lines)
        write_df_to_sheet(writer, "Afternoon Absent", df_afternoon, title_lines)

    output.seek(0)
    return send_file(
        output,
        download_name=f'Absent_Students_{today}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ---------- DOWNLOAD FULL ATTENDANCE WITH FILTER (Morning/Afternoon) ----------
@app.route('/download_full_attendance', methods=['POST'])
def download_full_attendance():
    filter_dept = request.form.get('filter_dept', '')
    filter_year = request.form.get('filter_year', '')

    columns = ['Date', 'Time', 'Register Number', 'Name', 'Department', 'Year']

    # Base query for a slot
    base_query = """
        SELECT a.date, TIME_FORMAT(a.time, '%%H:%%i:%%s'),
               s.register_number, s.name, s.department, s.year
        FROM attendance a
        JOIN students s ON a.student_id = s.id
        WHERE a.slot=%s
    """
    params_base = []

    if filter_dept:
        base_query += " AND s.department=%s"
        params_base.append(filter_dept)
    if filter_year:
        base_query += " AND s.year=%s"
        params_base.append(filter_year)
    base_query += " ORDER BY a.date DESC, a.time ASC, s.register_number ASC"

    try:
        with db_session(commit_on_success=False) as cur:
            # Morning attendance
            cur.execute(base_query, ['morning'] + params_base)
            morning_rows = cur.fetchall()

            # Afternoon attendance
            cur.execute(base_query, ['afternoon'] + params_base)
            afternoon_rows = cur.fetchall()
    except Exception as e:
        flash(f"❌ Download failed: {e}")
        return redirect(url_for('admin_dashboard'))

    df_morning = format_attendance_dataframe(morning_rows, columns)
    df_afternoon = format_attendance_dataframe(afternoon_rows, columns)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        title_lines = [
            "ADHIYAAMAAN COLLEGE OF ENGINEERING, HOSUR",
            f"Full Attendance - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        # Write Morning sheet
        write_df_to_sheet(writer, "Morning Attendance", df_morning, title_lines)
        # Write Afternoon sheet
        write_df_to_sheet(writer, "Afternoon Attendance", df_afternoon, title_lines)

    output.seek(0)
    return send_file(
        output,
        download_name=f'Full_Attendance_{datetime.now().strftime("%Y-%m-%d")}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )




# --- Student Registration with Auto-login & Success Page ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        data = request.form

        # --- Step 1: Capture & encode face ---
        frame = capture_image_from_base64(data['captured_image'])
        encoding, _ = encode_face_from_frame(frame)
        if encoding is None:
            flash("❌ Face not detected. Please try again.")
            return redirect('/register')

        face_encoding_str = serialize_encoding(encoding)

        # --- Step 2: Prepare folder & save image ---
        folder_path = f"static/student_photos/{data['department']}/{data['year']}"
        os.makedirs(folder_path, exist_ok=True)
        filename = f"{data['register_number']}.jpg"
        photo_path = os.path.join(folder_path, filename)
        save_image(frame, photo_path)

        # --- Step 3: Hash password ---
        hashed_password = generate_password_hash(data['password'])

        try:
            with db_session() as cur:
                # --- Step 4: Check for duplicate register number ---
                cur.execute("SELECT id FROM students WHERE register_number=%s", (data['register_number'],))
                existing = cur.fetchone()
                if existing:
                    flash("❌ Register number already exists!")
                    if os.path.exists(photo_path):
                        os.remove(photo_path)
                    return redirect('/register')

                # --- Step 5: Insert student ---
                cur.execute("""
                    INSERT INTO students
                    (name, register_number, course_type, department, year, batch, dob, mobile, blood_group, email, password, face_encoding, latitude, longitude, photo)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    data['name'], data['register_number'], data['course'], data['department'],
                    data['year'], data['batch'], data['dob'], data['mobile'], data.get('blood_group',''),
                    data['email'], hashed_password, face_encoding_str, data['latitude'], data['longitude'],
                    photo_path.replace("static/", "")
                ))

                # --- Step 6: Auto-login student ---
                cur.execute("SELECT * FROM students WHERE register_number=%s", (data['register_number'],))
                student = cur.fetchone()
                session['student_id'] = student['id']
                session['student_name'] = student['name']
                session['register_number'] = student['register_number']

            # --- Step 7: Show registration success page ---
            return render_template('registration_success.html', student_name=student['name'])

        except pymysql.err.IntegrityError:
            if os.path.exists(photo_path):
                os.remove(photo_path)
            flash("❌ Registration failed: duplicate or constraint error.")
            return redirect('/register')

        except Exception as e:
            if os.path.exists(photo_path):
                os.remove(photo_path)
            flash(f"❌ Registration failed: {e}")
            return redirect('/register')

    return render_template('register.html')


# --- student login path ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        reg = request.form['register_number']
        password = request.form['password']
        try:
            with db_session(commit_on_success=False) as cur:
                cur.execute("SELECT * FROM students WHERE register_number = %s", (reg,))
                student = cur.fetchone()
        except Exception as e:
            flash(f"DB error: {e}")
            student = None

        if not student:
            flash("❌ Register number not found")
        elif not check_password_hash(student['password'], password):
            flash("❌ Incorrect password")
        else:
            session['student_id'] = student['id']
            session['register_number'] = student['register_number']
            session['student_name'] = student['name']
            session['student_department'] = student['department']
            session['student_year'] = student['year']
            return redirect('/dashboard')

    return render_template('login.html')

# --- Forgot  student password request ---
@app.route('/student_forgot_password', methods=['GET', 'POST'])
def student_forgot_password():
    if request.method == 'POST':
        reg_no = request.form['register_number']
        email = request.form['email']

        try:
            with db_session() as cur:
                cur.execute("SELECT id FROM students WHERE register_number=%s AND email=%s", (reg_no, email))
                student = cur.fetchone()
                if student:
                    token = secrets.token_urlsafe(16)
                    cur.execute("UPDATE students SET reset_token=%s WHERE id=%s", (token, student['id']))
                    reset_link = url_for('student_reset_password', token=token, _external=True)
                    msg = Message("Password Reset Request", sender="noreply@example.com", recipients=[email])
                    msg.body = f"Click this link to reset your password: {reset_link}"
                    mail.send(msg)
                    flash("✅ Check your email for the reset link!")
                else:
                    flash("❌ Register number or email not found!")
        except Exception as e:
            flash(f"Email/DB error: {e}")
        return redirect(url_for('student_forgot_password'))

    return render_template('student_forgot_password.html')

# --- Reset  student password using token ---
@app.route('/student_reset_password/<token>', methods=['GET', 'POST'])
def student_reset_password(token):
    try:
        with db_session(commit_on_success=False) as cur:
            cur.execute("SELECT id FROM students WHERE reset_token=%s", (token,))
            student = cur.fetchone()
    except Exception as e:
        flash(f"DB error: {e}")
        return redirect(url_for('login'))

    if not student:
        flash("❌ Invalid or expired token!")
        return redirect(url_for('login'))

    if request.method == 'POST':
        new_password = request.form['new_password']
        import re
        pattern = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&]).{6,}$'
        if not re.match(pattern, new_password):
            flash("❌ Password must contain at least 6 characters, including uppercase, lowercase, number, and special character.")
            return redirect(request.url)

        hashed_pw = generate_password_hash(new_password)
        try:
            with db_session() as cur:
                cur.execute("UPDATE students SET password=%s, reset_token=NULL WHERE id=%s", (hashed_pw, student['id']))
            flash("✅ Password reset successfully! Login now.")
            return redirect(url_for('login'))
        except Exception as e:
            flash(f"❌ Password reset failed: {e}")
            return redirect(request.url)

    return render_template('student_reset_password.html')

# (Legacy) Reset by token route — fixed to use db_session
@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if request.method == 'POST':
        new_password = request.form['new_password']
        hashed_password = generate_password_hash(new_password)
        try:
            with db_session() as cur:
                cur.execute("UPDATE students SET password=%s WHERE reset_token=%s", (hashed_password, token))
            flash("✅ Password reset successfully!", "success")
            return redirect(url_for('login'))
        except Exception as e:
            flash(f"❌ Reset failed: {e}")
            return redirect(url_for('reset_password', token=token))
    return render_template('reset_password.html')

#student dashboard path
@app.route('/dashboard')
def dashboard():
    if 'student_id' not in session:
        return redirect('/login')

    student = None
    try:
        with db_session(commit_on_success=False) as cur:
            cur.execute("SELECT * FROM students WHERE id = %s", (session['student_id'],))
            student = cur.fetchone()

        if student and student.get('photo'):
            # Normalize path for Flask static folder
            student['photo'] = student['photo'].replace("\\", "/")  # Windows path fix

    except Exception as e:
        flash(f"DB error: {e}")
        student = None

    return render_template('student_dashboard.html', student=student)

# --- Student Mark Attendance ---
# --- Student Mark Attendance with Morning/Afternoon Slot ---
@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    if 'student_id' not in session:
        return redirect('/login')

    now = datetime.now().time()
# Morning: 08:30–10:30, Afternoon: 13:30–15:30
    if time(8, 30) <= now < time(10, 30):
        current_slot = "morning"
    elif time(13, 30) <= now < time(15, 30):
        current_slot = "afternoon"
    else:
        flash("⏰ Attendance can only be marked between 08:30–10:30 or 13:30–15:30.")
        session['attendance_marked_status'] = 'fail'
        return redirect('/attendance_marked_status')

    data = request.form
    frame = capture_image_from_base64(data['captured_image'])
    encoding, _ = encode_face_from_frame(frame)
    if encoding is None:
        flash("❌ Face not recognized. Try again.")
        session['attendance_marked_status'] = 'fail'
        return redirect('/attendance_marked_status')

    lat = float(data['latitude'])
    lon = float(data['longitude'])
    if not is_within_college_location(lat, lon, radius_km=50):
        flash("📍 Not within allowed location.")
        session['attendance_marked_status'] = 'fail'
        return redirect('/attendance_marked_status')

    try:
        with db_session(commit_on_success=False) as cur:
            cur.execute("SELECT * FROM students WHERE id = %s", (session['student_id'],))
            student = cur.fetchone()

        if not student:
            flash("❌ Student not found.")
            session['attendance_marked_status'] = 'fail'
            return redirect('/attendance_marked_status')

        with db_session(commit_on_success=False) as cur:
            cur.execute(
                "SELECT * FROM attendance WHERE student_id=%s AND date=CURDATE() AND slot=%s",
                (student['id'], current_slot)
            )
            if cur.fetchone():
                flash(f"❌ Attendance already marked for {current_slot}.")
                session['attendance_marked_status'] = 'fail'
                return redirect('/attendance_marked_status')

        known_encoding = deserialize_encoding(student['face_encoding'])
        if not compare_faces(known_encoding, encoding):
            flash("❌ Face does not match.")
            session['attendance_marked_status'] = 'fail'
            return redirect('/attendance_marked_status')

        # ----- Save attendance image -----
        today_date = datetime.now().date()  # YYYY-MM-DD as date object
        current_time = datetime.now().time()  # HH:MM:SS as time object

        folder_path = os.path.join("static", "attendance_images",
                                   student['department'], str(student['year']), str(today_date))
        os.makedirs(folder_path, exist_ok=True)
        filename = f"{student['register_number']}_{current_slot}_{today_date}.jpg"
        path = os.path.join(folder_path, filename)
        cv2.imwrite(path, frame)
        db_path = os.path.relpath(path, "static").replace("\\", "/")

        # ----- Insert attendance into DB -----
        with db_session() as cur:
            cur.execute("""
                INSERT INTO attendance
                (student_id, register_number, name, department, year, date, time, latitude, longitude, live_image, slot)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                student['id'], student['register_number'], student['name'],
                student['department'], student['year'], today_date, current_time,
                lat, lon, db_path, current_slot
            ))

        flash(f"✅ {current_slot.capitalize()} attendance marked successfully.")
        session['attendance_marked_status'] = 'success'
        return redirect('/attendance_marked_status')

    except Exception as e:
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass
        flash(f"❌ Attendance failed: {e}")
        session['attendance_marked_status'] = 'fail'
        return redirect('/attendance_marked_status')


@app.route('/attendance_marked_status')
def attendance_marked_status():
    status = session.get('attendance_marked_status', 'fail')
    return render_template('attendance_marked_status.html', attendance_status=status)

# ---------- STUDENT EXPORT ATTENDANCE (DATE RANGE + STATUS + LOCATION ) ----------


@app.route('/student_export_excel', methods=['GET', 'POST'])
def student_export_excel():
    if 'student_id' not in session:
        return redirect(url_for('login'))

    student_id = session['student_id']
    register_number = session.get('register_number')
    student_name = session.get('student_name', 'Unknown')

    if request.method == 'POST':
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')

        if not start_date or not end_date:
            flash("Please provide both start and end dates!", "danger")
            return redirect(url_for('student_dashboard'))

        conn = mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='ace_attendance_system'
        )
        cur = conn.cursor(dictionary=True)

        # Fetch attendance records
        cur.execute("""
            SELECT date, slot, time, latitude, status 
            FROM attendance 
            WHERE register_number = %s AND date BETWEEN %s AND %s
            ORDER BY date ASC
        """, (register_number, start_date, end_date))
        records = cur.fetchall()
        conn.close()

        # Convert to DataFrame
        df = pd.DataFrame(records)

        # Generate all dates between start & end
        all_dates = pd.date_range(start=start_date, end=end_date)
        data = []

        for date in all_dates:
            day = date.strftime("%A")
            morning = next((r for r in records if r['date'] == date.date() and r['slot'] == 'morning'), None)
            afternoon = next((r for r in records if r['date'] == date.date() and r['slot'] == 'afternoon'), None)

            data.append({
                'Date': date.strftime("%d %b %Y"),
                'Day': day,
                'Morning Status': morning['status'] if morning else 'absent',
                'Morning Time': morning['time'] if morning else '-',
                'Morning Latitude': morning['latitude'] if morning else '-',
                'Afternoon Status': afternoon['status'] if afternoon else 'absent',
                'Afternoon Time': afternoon['time'] if afternoon else '-',
                'Afternoon Latitude': afternoon['latitude'] if afternoon else '-'
            })

        df_full = pd.DataFrame(data)

        # Create Excel
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = f"{student_name}_Attendance"

        headers = list(df_full.columns)
        ws.append(headers)

        for index, row in df_full.iterrows():
            ws.append(list(row.values))

        # Apply styling
        for row in ws.iter_rows(min_row=2, min_col=1):
            date_cell = row[0]
            day_cell = row[1]
            day_name = day_cell.value
            morning_status = row[2]
            afternoon_status = row[5]

            # Sunday orange
            if day_name == "Sunday":
                for cell in row:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

            # Saturday blue
            elif day_name == "Saturday":
                for cell in row:
                    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            # Status colors
            if morning_status.value == 'present':
                morning_status.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                morning_status.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

            if afternoon_status.value == 'present':
                afternoon_status.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                afternoon_status.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

        # Header style
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 3

        wb.save(output)
        output.seek(0)

        filename = f"{student_name}_Attendance_{start_date}_to_{end_date}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return redirect(url_for('student_dashboard'))


# --- Cleanup old attendance images (>100 days) ---
def cleanup_old_images():
    base_path = os.path.join("static", "attendance_images")
    cutoff_date = datetime.now() - timedelta(days=100)

    for root, dirs, files in os.walk(base_path):
        for file in files:
            if file.lower().endswith(".jpg"):
                file_path = os.path.join(root, file)
                try:
                    file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                    if file_time < cutoff_date:
                        os.remove(file_path)
                        print(f"🗑️ Deleted old file: {file_path}")
                except Exception as e:
                    print(f"⚠️ Error deleting {file_path}: {e}")

    # --- Remove empty folders after cleanup ---
    for root, dirs, files in os.walk(base_path, topdown=False):
        for dir in dirs:
            dir_path = os.path.join(root, dir)
            if not os.listdir(dir_path):  # if folder is empty
                os.rmdir(dir_path)
                print(f"📂 Removed empty folder: {dir_path}")


# --- Scheduler setup ---
scheduler = BackgroundScheduler()
scheduler.add_job(func=cleanup_old_images, trigger="interval", days=1)  # run daily
scheduler.start()

# Ensure scheduler shuts down on app exit
atexit.register(lambda: scheduler.shutdown())



@app.route('/test_email')
def test_email():
    try:
        msg = Message("Test Email", sender="naveenn13032004@gmail.com", recipients=["your_other_email@gmail.com"])
        msg.body = "✅ Flask email setup is working!"
        mail.send(msg)
        return "Email sent successfully!"
    except Exception as e:
        return f"Error: {e}"
    
    # --- Test Database Data ---
@app.route('/test_data')
def test_data():
    today = date.today()
    
    with db_session() as cur:
        # Check students
        cur.execute("SELECT id, register_number, name FROM students")
        students = cur.fetchall()
        
        # Check today's attendance
        cur.execute("SELECT student_id, slot FROM attendance WHERE date = %s", (today,))
        attendance = cur.fetchall()
        
        # Check morning absent manually
        cur.execute("""
            SELECT s.id, s.register_number, s.name 
            FROM students s 
            WHERE s.id NOT IN (
                SELECT student_id FROM attendance 
                WHERE date = %s AND slot = 'morning'
            )
        """, (today,))
        manual_morning_absent = cur.fetchall()
        
        # Check afternoon absent manually
        cur.execute("""
            SELECT s.id, s.register_number, s.name 
            FROM students s 
            WHERE s.id NOT IN (
                SELECT student_id FROM attendance 
                WHERE date = %s AND slot = 'afternoon'
            )
        """, (today,))
        manual_afternoon_absent = cur.fetchall()
    
    return f"""
    <h3>Database Test - Date: {today}</h3>
    
    <h4>All Students ({len(students)}):</h4>
    <pre>{students}</pre>
    
    <h4>Today's Attendance ({len(attendance)}):</h4>
    <pre>{attendance}</pre>
    
    <h4>Manual Morning Absent ({len(manual_morning_absent)}):</h4>
    <pre>{manual_morning_absent}</pre>
    
    <h4>Manual Afternoon Absent ({len(manual_afternoon_absent)}):</h4>
    <pre>{manual_afternoon_absent}</pre>
    """


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)